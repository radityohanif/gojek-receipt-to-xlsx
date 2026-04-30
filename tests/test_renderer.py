from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from gojek_receipt.core.models import Receipt, Transaction
from gojek_receipt.core.renderer import render


def _sample_receipt() -> Receipt:
    return Receipt(
        nama="Budi Santoso",
        periode="01/04/2026 - 29/04/2026",
        total_transaksi=2,
        transactions=[
            Transaction(
                tanggal=date(2026, 4, 29),
                waktu=time(10, 2, 32),
                no_transaksi="RB-123456-789",
                layanan="GoRide",
                dari="Sudirman",
                tujuan="Kuningan",
                metode_bayar="GoPay",
                total_dibayar=15000,
            ),
            Transaction(
                tanggal=date(2026, 4, 28),
                waktu=time(14, 0, 0),
                no_transaksi="F-987654",
                layanan="GoFood",
                dari="Warung Nasi ABC",
                tujuan="Kuningan Barat",
                metode_bayar="Kartu kredit/debit",
                total_dibayar=75000,
            ),
        ],
    )


class TestRenderer:
    def test_renderer_creates_xlsx(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        assert path.exists()

    def test_renderer_headers(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "Transaksi"
        # Table header now at row 7
        assert ws.cell(7, 1).value == "Tanggal"
        assert ws.cell(7, 8).value == "Total Dibayar"
        assert ws.cell(7, 3).value == "No. Transaksi"

    def test_renderer_sum_formula(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        wb = load_workbook(path, data_only=False)
        ws = wb.active
        # Footer is row 10 (rows 1-6 header section, row 7 table header, rows 8-9 data, row 10 footer)
        footer_cell = ws.cell(10, 8).value or ""
        assert "SUM" in footer_cell

    def test_renderer_data_values(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Data now starts at row 8
        assert ws.cell(8, 3).value == "RB-123456-789"
        assert ws.cell(8, 8).value == 15000
        assert ws.cell(9, 4).value == "GoFood"

    def test_renderer_number_format(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        wb = load_workbook(path)
        ws = wb.active
        # Check that Total Dibayar has number format (first data row is now row 8)
        cell_h8 = ws.cell(8, 8)
        assert cell_h8.number_format == "#,##0"

    def test_renderer_empty_receipt(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        empty = Receipt(
            nama="Test User", periode="01-29/04/2026", total_transaksi=0, transactions=[]
        )
        render(empty, path)
        wb = load_workbook(path)
        ws = wb.active
        # Table header at row 7, TOTAL at row 8
        assert ws.cell(7, 1).value == "Tanggal"
        assert ws.cell(8, 1).value == "TOTAL"

    def test_renderer_header_section(self, tmp_path: Path) -> None:
        path = tmp_path / "out.xlsx"
        render(_sample_receipt(), path)
        wb = load_workbook(path)
        ws = wb.active
        # Check header section content
        assert ws.cell(1, 2).value == "Riwayat Transaksi Gojek"
        assert ws.cell(3, 1).value == "Periode:"
        assert ws.cell(3, 2).value == "01/04/2026 - 29/04/2026"
        assert ws.cell(4, 1).value == "Pemilik:"
        assert ws.cell(4, 2).value == "Budi Santoso"
        assert ws.cell(5, 1).value == "Total Transaksi:"
        assert ws.cell(5, 2).value == 2
