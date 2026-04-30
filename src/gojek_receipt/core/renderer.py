from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from gojek_receipt.core.models import Receipt, Transaction


# ─── Style constants ──────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="16A34A")

_FILL_GORIDE = PatternFill(fill_type="solid", fgColor="DCFCE7")
_FILL_GOFOOD = PatternFill(fill_type="solid", fgColor="FEF3C7")
_FILL_GOTRANSIT = PatternFill(fill_type="solid", fgColor="DBEAFE")
_FILL_DEFAULT = PatternFill(fill_type="solid", fgColor="FFFFFF")

_FOOTER_FONT = Font(bold=True)
_FOOTER_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")

_HEADERS = [
    "Tanggal",
    "Waktu",
    "No. Transaksi",
    "Layanan",
    "Dari",
    "Tujuan",
    "Metode Bayar",
    "Total Dibayar",
]

_COL_WIDTHS = [12, 12, 22, 18, 28, 28, 20, 16]

_DATE_FMT = "DD/MM/YYYY"
_TIME_FMT = "HH:MM:SS"
_AMOUNT_FMT = "#,##0"


def _row_fill(layanan: str) -> PatternFill:
    """Determine row color based on service type."""
    s = layanan.lower()
    if "goride" in s or "gride" in s:
        return _FILL_GORIDE
    if "gofood" in s or "food" in s:
        return _FILL_GOFOOD
    if "gotransit" in s or "transit" in s:
        return _FILL_GOTRANSIT
    return _FILL_DEFAULT


def _ensure_parent(path: Path) -> None:
    """Ensure parent directory exists."""
    path.parent.mkdir(parents=True, exist_ok=True)


def _add_header_section(ws, receipt: Receipt) -> None:
    """Add header section with logo, title, period, and owner info."""
    # Set row heights for header section
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 5
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 5

    # Try to add logo (if exists)
    logo_path = Path(__file__).parent.parent.parent / "logo.png"
    if logo_path.exists():
        try:
            img = Image(str(logo_path))
            img.width = 30
            img.height = 30
            ws.add_image(img, "A1")
        except Exception:
            pass

    # Row 1: Title + Logo
    title_cell = ws.cell(1, 2, "Riwayat Transaksi Gojek")
    title_cell.font = Font(bold=True, size=16, color="16A34A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Info label styling
    header_info_font = Font(bold=True, size=11, color="FFFFFF")
    header_info_fill = PatternFill(fill_type="solid", fgColor="16A34A")
    header_info_alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: Period info
    period_label = ws.cell(3, 1, "Periode:")
    period_label.font = Font(bold=True, size=11)
    period_label.alignment = header_info_alignment

    period_value = ws.cell(3, 2, receipt.periode)
    period_value.font = Font(size=11)
    period_value.alignment = header_info_alignment

    # Row 4: Owner name
    name_label = ws.cell(4, 1, "Pemilik:")
    name_label.font = Font(bold=True, size=11)
    name_label.alignment = header_info_alignment

    name_value = ws.cell(4, 2, receipt.nama)
    name_value.font = Font(size=11)
    name_value.alignment = header_info_alignment

    # Row 5: Total transactions
    total_label = ws.cell(5, 1, "Total Transaksi:")
    total_label.font = Font(bold=True, size=11)
    total_label.alignment = header_info_alignment

    total_value = ws.cell(5, 2, receipt.total_transaksi)
    total_value.font = Font(size=11)
    total_value.alignment = header_info_alignment


def render(receipt: Receipt, path: Path) -> Path:
    """Write receipt data to an XLSX workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"

    # ─── Add logo and header info ──────────────────────────────────────────────
    _add_header_section(ws, receipt)

    # ─── Transaction table header row (row 7) ─────────────────────────────────
    table_header_row = 7
    for col, label in enumerate(_HEADERS, start=1):
        cell = ws.cell(table_header_row, col, label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    # ─── Data rows (start from row 8) ──────────────────────────────────────────
    for r_idx, tx in enumerate(receipt.transactions, start=8):
        fill = _row_fill(tx.layanan)
        _write_row(ws, r_idx, tx, fill)

    # ─── Footer SUM row ───────────────────────────────────────────────────────
    footer_row = len(receipt.transactions) + 8
    last_data_row = footer_row - 1

    ws.cell(footer_row, 1, "TOTAL").font = _FOOTER_FONT
    ws.cell(footer_row, 1).fill = _FOOTER_FILL
    ws.cell(footer_row, 1).border = _BORDER

    # SUM formula for Total Dibayar (column H = 8, data starts at row 8)
    sum_cell = ws.cell(footer_row, 8, f"=SUM(H8:H{last_data_row})")
    sum_cell.number_format = _AMOUNT_FMT
    sum_cell.font = _FOOTER_FONT
    sum_cell.fill = _FOOTER_FILL
    sum_cell.border = _BORDER

    # Fill remaining footer cells
    for col in range(2, 8):
        c = ws.cell(footer_row, col)
        c.fill = _FOOTER_FILL
        c.border = _BORDER

    # ─── Column widths ────────────────────────────────────────────────────────
    col_letters = "ABCDEFGH"
    for i, width in enumerate(_COL_WIDTHS):
        ws.column_dimensions[col_letters[i]].width = width

    # Freeze table header row
    ws.freeze_panes = "A8"

    _ensure_parent(path)
    wb.save(path)
    return path


def _write_row(ws, row: int, tx: Transaction, fill: PatternFill) -> None:
    """Write a single transaction row to the worksheet."""
    cells_data = [
        (1, tx.tanggal, _DATE_FMT),
        (2, tx.waktu, _TIME_FMT),
        (3, tx.no_transaksi, None),
        (4, tx.layanan, None),
        (5, tx.dari, None),
        (6, tx.tujuan, None),
        (7, tx.metode_bayar, None),
        (8, tx.total_dibayar, _AMOUNT_FMT),
    ]
    for col, value, fmt in cells_data:
        cell = ws.cell(row, col, value)
        cell.fill = fill
        cell.border = _BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        if fmt:
            cell.number_format = fmt
